#open excel list of entry and glosses in Kamus Dewan
from openpyxl import load_workbook
wb = load_workbook('Kamus Dewan.xlsx')
ws = wb['Sheet1']

#open list of pos terms 
f = open("bahasa_pos_pruned.txt", "r")
pos_terms = f.read()
pos_terms = pos_terms.split("\n")
f.close()

#open list of neg terms 
f1 = open("bahasa_neg_pruned.txt", "r")
neg_terms = f1.read()
neg_terms = neg_terms.split("\n")
f1.close()

pos_matched_entries = []
neg_matched_entries = []

#find pos_term matches within gloss_terms and retrieve gloss entry   
counter = 0    
for x in range(1,55725):
    counter += 1
    entry = ws.cell(row = x, column = 3)
    entry = entry.value
    entry = entry.split(" ")
    entry_term = entry[0]
    entry_gloss = entry[1:]
    print("---Entry number:", counter)
    print("Entry term:", entry_term)
    print("Entry_gloss:", entry_gloss)

#remove term after "bkn" to avoid mismatches 
    for i, gloss_term in enumerate(entry_gloss):
        if gloss_term.lower() == "bkn":
            entry_gloss[i+1] = ""

    pos_matches = 0
    neg_matches = 0
#get entries with matches
    for gloss_term in entry_gloss:
        gloss_term = gloss_term.lower() #normalize
        
        if (gloss_term == "tidak" or gloss_term == "tak" or gloss_term == "bukan" or gloss_term == "jangan" or gloss_term == "belum"):
            break #break on negation matches to handle negation
            
        if entry_term == "~": #handle the ~ appearing at start of some glosses
            entry_term = entry[1]
            entry_gloss = entry[2:]
            
        for pos_term in pos_terms: # increment pos_matches when any gloss term matches with pos seed
            if pos_term == gloss_term and gloss_term != "":
                pos_matches += 1
                print("pos match " + str(pos_matches) + ": " + pos_term + " --> in gloss of " + entry_term + " in row " + str(x))

        for neg_term in neg_terms: # increment neg_matches
            if neg_term == gloss_term and gloss_term != "":
                neg_matches += 1
                print("neg match " + str(neg_matches) + ": " + neg_term + " --> in gloss of " + entry_term + " in row " + str(x))

    print("final pos_match_count: ", pos_matches)
    print("final neg_match_count: ", neg_matches)
    
    if pos_matches > neg_matches: #final comparison between pos_matches and neg_matches
        pos_matched_entries += [entry_term]
        print(entry_term + " matched as POS")
    elif neg_matches > pos_matches:
        neg_matched_entries += [entry_term]
        print(entry_term + " matched as NEG")
    else:
        print("entry OBJ and discarded")
        
    print("\n")

print("pos_matched entries count:", len(set(pos_matched_entries)))    
#print(pos_matched_entries)

#write final pos results to txt
f2 = open("pos output from kd.txt", "a", encoding = 'utf-8')
f2.write("\npos matched entries:\n")
f2.write(str(set(pos_matched_entries)))
f2.close()

print("neg_matched entries count:", len(set(neg_matched_entries)))    
#print(neg_matched_entries)

#write final neg results to txt
f3 = open("neg output from kd.txt", "a", encoding = 'utf-8')
f3.write("\nneg matched entries:\n")
f3.write(str(set(neg_matched_entries)))
f3.close()